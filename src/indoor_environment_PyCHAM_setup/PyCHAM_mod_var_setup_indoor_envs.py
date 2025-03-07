'''code to write the model inputs for simulations
of particulate matter for indoor environments'''
# the resulting model variables file and excel 
# file containing the influx
# rates of certain chemicals in certain phases 
# can be used as input to the CHemistry with
# Aerosol Microphysics in Python (PyCHAM) model:
# https://github.com/simonom/PyCHAM

# import depdencies
import numpy as np
import os
import platform
import openpyxl
import ast
import scipy.constants as si
import math

# user-defined variables start --------------------------------

# set base path - to be used for prefixing other paths (where wanted)
base_path = str('/Users/user/Library/CloudStorage/' +
			'OneDrive-TheUniversityofManchester')

# path to folder where model variable and continuous influx inputs will be saved
path_there = str(base_path + '/INGENIOUS/Papers/' +
		'simulated_PM_mass_versus_observation/PyCHAM_in/')

# path to chemical scheme file
chem_sch_name = str(base_path + '/INGENIOUS/Papers/' +
		'simulated_PM_mass_versus_observation/PyCHAM_in/' +
		'fullMCM_PRAMAP_autoAPRAMBZ.kpp')

# path to xml file
xml_name = str(base_path + '/INGENIOUS/Papers/' +
		'simulated_PM_mass_versus_observation/PyCHAM_in/' +
		'MCM_PRAM_autoAPRAMfw_xml.xml')

# path to results folder
res_path = str(base_path + '/INGENIOUS/Papers/' +
		'simulated_PM_mass_versus_observation/PyCHAM_out/')


# the volume (cm3) of the envelope that emissions are released into
# and therefore diluted
env_vol = (5.e2*6.e2*6.e2)

# user-defined variables end --------------------------------

# define function
def mod_var_setup(path_there, chem_sch_name, xml_name, res_path, base_path):

	# function creates the model variable input file
	# for home simulations for PyCHAM for the INGENIOUS
	# project. For the particle-phase, the code 
	# finds the mean radius of particles per size bin
	# by first calculating the total volume of
	# particles per size bin and then dividing by
	# number concentration of particles and then
	# converting volume to radius. For the 
	# particle-phase concentration per component, mole
	# fractions of components from the outdoor source
	# and from the indoor activity source are weighted 
	# by the total molecular concentration in each 
	# source, then the resulting mole fractions are 
	# used to estimate the mean density and molar mass
	# of particle-phase components, so that a mean
	# molar volume can be estimated, and this is
	# combined with total volume to estimate rate of
	# influx of molecular concentrations of components 
	# (molecules/cm3/s)

	# inputs: ----------------------
	# path_there - path to save model variable files to
	# chem_sch_name - path to chemical scheme to use
	# xml_name - path to xml file to use
	# res_path - path to results folder
	# base_path - path to INGENIOUS folder
	# ------------------------------

	# set range for air change rate (for maximum:
	# https://www.vent-axia.com/sites/default/files/
	# 2023-11/97ee4be4-7c33-4f0d-a775-5b8c83ab1f5e.pdf)
	acr_range = [1.]

	# set time for acr (s through simulation)
	acr_times = np.array((0.)).reshape(1)

	# note that inside photolysisRates is a prescribed
	# wavelength-dependent transmission factor
	# applied to the Hayman estimation of J values
	# based on Figure 1b of 
	# doi.org/10.1007/s44223-022-00009-6, with further 
	# explanation in photolysisRates
	light_status = '3'		

	
	# set range for tranmission factor 
	# of light through glass, to be used in addition
	# to the wavelength-dependent transmission factor
	# in photolysisRates
	tf_list = [0.5]

	# set activity frequency
	mop_freqs = [0]
	fry_freqs = [0]
	dust_freqs = [0]
	bath_freqs = [0]
	pcp_freqs = [0]
	frag_freqs = [0]
	wbs_freqs = [0]

	# set activity start times (hours through day)
	mop_times = []
	fry_times = []
	dust_times = []
	bath_times = []
	pcp_times = []
	frag_times = []
	wbs_times = []

	# set activity duration (hours taken)
	mop_dur = 0.33
	fry_dur = 0.33
	dust_dur = 0.33
	bath_dur = 0.25
	# 1 second, which, when combined with the emission rate for 
	# the activity, gives a body spray total mass equivalent of 
	# 1.5 g (which is typical of a spraying session, as 
	# detailed in the corresponding column header of 
	# indoor_emi_pri_org_VBS)
	pcp_dur = 2.78e-4
	frag_dur = 1.
	wbs_dur = 1.5

	# outdoor concentration names (first letter for season (s for summer), 
	# second for meterological condition (g for stagnant and dry), third 
	# for location (r for rural))
	oc_rn = ['sgu']

	# loading emission rates from indoor activities start ------------
	# get the names of all components with influx from
	# indoor activity
	wb = openpyxl.load_workbook(filename = str(base_path + 
		'/INGENIOUS/mod_var_setup/' +
		'indoor_emi_pri_org_VBS.xlsx'))

	# get active sheet
	sheet = wb['indoor_emi']
	
	ir = 0 # count on rows

	# prepare to hold index of column with component names
	comp_name_col_indx = []
	# prepare to hold title of column with component names
	comp_name_col_head = []
	# prepare to hold index of columns with influx rates
	comp_inf_col_indx = []
	# prepare to hold title of column with influx rates
	comp_inf_col_head = []
	# prepare to hold component names
	ind_emi_comp_names0 = np.zeros((0))
	# prepare to hold indoor emission values
	ind_emi_val = np.zeros((0, 0))

	# flag to start collecting influx rates
	inf_col = 0

	# loop through rows
	for i in sheet.iter_rows(values_only = True):

		# if at the end of rows with values
		if (sum(np.array((i)) != None) == 0):
			break

		ir += 1 # count on rows
		
		# if on row containing column headings
		if ('MCM Name' in str(i)):
		
			# flag to start collecting influx rates
			inf_col = 1

			# count on columns
			ic = -1
			
			# loop through column headers
			for col_head in i:

				# count on columns
				ic += 1
				
				if 'MCM Name' in col_head:
					comp_name_col_indx.append(ic)
					comp_name_col_head.append('comp_name')
				if 'baseline' in col_head:
					comp_inf_col_indx.append(ic)
					comp_inf_col_head.append('non')
				if 'igarette' in col_head:
					comp_inf_col_indx.append(ic)
					comp_inf_col_head.append('cigs')
				if 'Frying' in col_head or 'frying' in col_head:
					comp_inf_col_indx.append(ic)
					comp_inf_col_head.append('fry')
				if 'Bleach' in col_head or 'bleach' in col_head:
					comp_inf_col_indx.append(ic)
					comp_inf_col_head.append('bop')
				if 'natural' in col_head and 'mop' in col_head:
					comp_inf_col_indx.append(ic)
					comp_inf_col_head.append('mop')
				if 'wood' in col_head or 'Wood' in col_head:
					comp_inf_col_indx.append(ic)
					comp_inf_col_head.append('wst')
				if 'shower' in col_head or 'Shower' in col_head:
					comp_inf_col_indx.append(ic)
					comp_inf_col_head.append('sho')
				if 'pcp' in col_head:
					comp_inf_col_indx.append(ic)
					comp_inf_col_head.append('pcp')
			continue # onto next row
		
		if (inf_col == 1):
			ind_emi_comp_names0 = np.concatenate((
			ind_emi_comp_names0, 
			np.array((i))[comp_name_col_indx]), axis=0)
			# setup matrix with components (both gas- and 
			# particle-phase) in rows and
			# activities in columns
			if (ind_emi_val.shape[0] == 0):
				ind_emi_val = np.zeros((1, len(comp_inf_col_indx)))
			else:
				ind_emi_val = np.concatenate((ind_emi_val, 
				np.zeros((1, len(comp_inf_col_indx)))), axis=0)
			
			# contain spreadsheet values in array
			ind_emi_val[-1, :] = np.array((i))[comp_inf_col_indx]

	# loading emission rates from indoor activities end --------------
	
	# estimate total number of simulations
	num_sim = (len(acr_range)*len(tf_list)*len(mop_freqs)*len(fry_freqs)*
		len(bath_freqs)*len(pcp_freqs)*len(wbs_freqs))

	# loop through simulations
	for simi in range(num_sim):

		# initiate results folder name
		res_nam = ''

		# setting parameter values starts -----------------

		# set outdoor condition
		oci = oc_rn[0]
		res_nam = str(res_nam + oci + '_')
	
		# set air change rate
		acr = np.array((acr_range[0])).reshape(1)

		# update results name
		res_nam = str(res_nam + str(acr[0]) + 'acr_')

		# set light transmission factor
		tf = tf_list[0]
		tf_str = str(tf)

		# update results name
		res_nam = str(res_nam + str(tf) + 'tf_')

		# set mop frequency
		mop_freq = mop_freqs[0]
		# update results name
		res_nam = str(res_nam + str(mop_freq) + 'mops_')

		# set fry frequency
		fry_freq = fry_freqs[0]
		# update results name
		res_nam = str(res_nam + str(fry_freq) + 'fry_')

		# set bath frequency
		bath_freq = bath_freqs[0]
		# update results name
		res_nam = str(res_nam + str(bath_freq) + 'sho_')

		# set personal care product frequency
		pcp_freq = pcp_freqs[0]
		res_nam = str(res_nam + str(pcp_freq) + 'pcp_')

		# set wood burning stove frequency
		wbs_freq = wbs_freqs[0]
		# update results name
		res_nam = str(res_nam + str(wbs_freq) + 'wst')

		# setting parameter values ends --------------------

		# setting model variables file starts ---------------

		# convert air change rate from fraction per 
		# hour to fraction per second
		acr = np.array((acr))/3.6e3

		# name specific path for saving these inputs to
		path_there_spec = str(path_there + res_nam)

		# create folder for this input
		os.makedirs(path_there_spec, exist_ok = True)

		# include directory separator in path name
		path_there_spec = str(path_there_spec + '/')

		# name and path for model variables file
		mod_var_path = str(path_there_spec + res_nam + '_mod_var.txt')

		# name and path for continuous influx file
		cont_infl_path = str(path_there_spec + res_nam + '_cont_infl.xlsx')
		
		# for model variable file,
		# w+ means that if file doesn't exist, a new file is created
		fo = open(mod_var_path, 'w+')

		# get each line as a subsequent item in a list
		lines = fo.readlines()

		# preceed results folder name with path to results folder
		res_nam_mod_var_file = str(res_path + res_nam)

		# remember path to original simulation results
		if (simi == 0):
			res_nam_orig = res_nam_mod_var_file
		
		lines.append(str('res_file_name = ' + res_nam_mod_var_file + '\n'))
		lines.append(str('total_model_time = ' + '8.64e4' + '\n'))
		lines.append(str('update_step = ' + '3.0e2' + '\n'))
		lines.append(str('recording_time_step = ' + '6.0e2' + '\n'))
		lines.append(str('lat = ' + '53.' + '\n'))
		lines.append(str('lon = ' + '1.' + '\n'))
		lines.append(str('temperature = ' + '293.15' + '\n'))
		lines.append(str('rh = ' + '0.60' + '\n'))
		lines.append(str('light_status = ' + light_status + '\n'))
		lines.append(str('trans_fac = ' + tf_str + '\n'))
		lines.append(str('p_init = ' + '101300' + '\n'))
		lines.append(str('number_size_bins = ' + '2' + '\n'))
		lines.append(str('Vwat_inc = 0' + '\n'))
		lines.append(str('seed_eq_wat = 0' + '\n'))
		lines.append(str('space_mode = man' + '\n'))
		lines.append(str('lower_part_size = 9.e-4, 1.25, 10.' + '\n'))
		lines.append(str('upper_part_size = 10.' + '\n'))
		lines.append(str('vol_Comp = bc, AMM_SUL, pri_org, ' +
		'sec_org-2, sec_org-1, sec_org0, sec_org1, bcin, ' +
		'pri_orgin, NO2_wall1, O3_wall1' + '\n'))
		lines.append(str('volP = 0.0, 0.0, 0.0, 8.12e-16, 9.75e-12, '+
		'1.22e-7, 1.62e-4, 0., 0., 0., 0.' + '\n'))
		lines.append(str('nonHOMs_vp_method = EVAPORATION\n'))
		lines.append(str('HOMs_vp_method = EVAPORATION\n'))
		lines.append(str('inorg_part_flag = 1\n'))
		lines.append(str('chem_scheme_markers = {, RO2, +, C(ind_, ), ' +
			', &, , [, :, }, ;,' + '\n'))
		lines.append(str('chem_sch_name = ' +  chem_sch_name + '\n'))
		lines.append(str('xml_name = ' +  xml_name + '\n'))
		# only do spin_up for first simulation
		if (simi == 0):
			lines.append(str('spin_up = ' + '2' + '\n'))
		else:
			lines.append(str('C0 = ' + res_nam_orig + '\n'))
		lines.append(str('pars_skip = 0' + '\n'))
		lines.append(str('dil_fac = ' + str(acr).replace(
		'\n', '').replace('[', '').replace(']', '').replace(
		' ', ',') + '\n'))
		lines.append(str('dil_fact = ' + str(acr_times).replace(
		'\n', '').replace('[', '').replace(']', '').replace(
		'. ', '.,').replace(' ', '') + '\n'))
		lines.append(str('daytime_start = ' + '0.' + '\n'))
		lines.append(str('cont_infl = ' + cont_infl_path + '\n'))

		lines.append(str('wall_on = ' + '1' + '\n'))
		lines.append(str('number_wall_bins = ' + '1' + '\n'))
		lines.append(str('coag_on = ' + '1' + '\n'))
		lines.append(str('McMurry_flag = ' + '0' + '\n'))
		lines.append(str('# Note that the deposition rate of particles '+
		'is taken from the PyCHAM inputs for comparison against the ' +
		'Tran et a. 2017 paper, used for EAC2023 results: ' +
		'INGENIOUS/Meetings/EAC2023/EAC2023_poster_input_output/Tran2017' + '\n'))
		lines.append(str('inflectDp = ' + '8.e-7' + '\n'))
		lines.append(str('Grad_pre_inflect = ' + '1.6' + '\n'))
		lines.append(str('Grad_post_inflect = ' + '0.6' + '\n'))
		lines.append(str('Rate_at_inflect = ' + '2.33e-5' + '\n'))
		lines.append(str('mass_trans_coeff = 2.7e-3; NO2_wall1_1.1e-4' + '\n'))
		lines.append(str('eff_abs_wall_massC = 1.e1' + '\n'))
		lines.append(str('ppartit_cutoff = 1.e2' + '\n'))
		lines.append(str('wpartit_cutoff = 1.e10' + '\n'))
		lines.append(str('z_prt_coeff_loC = 1.e-6' + '\n'))

		# set day of year, omit final \n as this is the final line
		if (res_nam[0] == 's'):
			lines.append(str('DayOfYear = ' + '183'))
		if (res_nam[0] == 'w'):
			lines.append(str('DayOfYear = ' + '1'))

		# write out updated text
		for line in lines:
			fo.write(line)
		# save and close file
		fo.close()

		# setting model variables file ends -----------------

		# start of continuous influx section ---------------
		
		# activity times and type matrix (activities in rows and times
		# (every 5 minutes in columns))
		act_matrix_act_order = np.array(('mop', 'fry', 'sho', 'pcp', 'wst'))
		act_matrix = np.zeros((len(act_matrix_act_order)+1, int((24*3600)/300)))
		act_matrix[0, :] = np.arange(0., (24.*3600.), 300.)
		
	
		# all times in hours
		all_times_hr = act_matrix[0, :]/3600.

		# activity times
		for mopi in mop_times[0:mop_freq]: # loop through times
			mop_ti = (all_times_hr >= mopi)*(all_times_hr < mopi+mop_dur)
			act_matrix[1, :][mop_ti] = 1
	
		for fryi in fry_times[0:fry_freq]: # loop through times
			fry_ti = (all_times_hr >= fryi)*(all_times_hr < fryi+fry_dur)
			act_matrix[2, :][fry_ti] = 1

		for bathi in bath_times[0:bath_freq]: # loop through times
			bath_ti = (all_times_hr >= bathi)*(all_times_hr < bathi+bath_dur)
			act_matrix[3, :][bath_ti] = 1

		for pcpi in pcp_times[0:pcp_freq]: # loop through times
			pcp_ti = (all_times_hr >= pcpi)*(all_times_hr < pcpi+pcp_dur)
			act_matrix[4, :][pcp_ti] = 1

		for wbsi in wbs_times[0:wbs_freq]: # loop through times
			wbs_ti = (all_times_hr >= wbsi)*(all_times_hr < wbsi+wbs_dur)
			act_matrix[5, :][wbs_ti] = 1

		# file name
		outd_dir_name = str('outprep_' + oci)

		# path to directory containing outdoor concentrations
		outd_dir_basic = str(base_path + '/INGENIOUS/'+
		'Meetings/EAC2024/PyCHAM_inputs/outd_concs/')

		# set path to directory
		outd_dir = str(outd_dir_basic + outd_dir_name)

		# withdraw times (s)
		fname = str(outd_dir + '/time')
		t_array = np.loadtxt(fname, delimiter=',', skiprows=1)

		# component names from outdoor concentration run
		load_path = str(outd_dir + '/comp_namelist.npy') # path
		comp_names = (np.load(load_path, allow_pickle=True)).tolist()

		# component molar masses (g/mol)
		load_path = str(outd_dir + '/y_mw.npy') # path
		y_MM = np.load(load_path, allow_pickle=True)

		# component molar volumes
		# cm3/mol (needed to get component densities below)
		load_path = str(outd_dir + '/MV.npy') # path
		MV = np.load(load_path, allow_pickle=True)

		# estimated component densities (g/cm3)
		dens = y_MM/MV

		# number of components from outdoor simulation
		nc = int(len(comp_names))
		# remember number of original components in outdoor
		# result
		nc0 = nc

		# get factor to multiply ppb to get molecules/cm3
		fname = str(outd_dir + '/model_and_component_constants')
		const_in = open(fname)
		for line in const_in.readlines():
			if (str(line.split(',')[0]) == 
				'factor_for_multiplying_ppb_to_get_molec/cm3_with_time'):

				# find index of first [ and index of last ]
				icnt = 0 # count on characters
				for i in line:
					if i == '[':
						st_indx = icnt
						break
					icnt += 1 # count on characters
				for cnt in range(10):
					if line[-cnt] == ']':
						fi_indx = -cnt+1
						break

				# conversion factor to change gas-phase 
				# concentrations from # molecules/cm3 
				# (air) into ppb
				Cfactor = ast.literal_eval(line[st_indx:fi_indx])

		# component SMILES
		load_path = str(outd_dir + '/rel_SMILES.npy') # path
		rel_SMILES = (np.load(load_path, allow_pickle=True)).tolist()

		# outdoor particle radii (um)
		fname = str(outd_dir + '/size_bin_radius')
		# skiprows=1 omits header
		x = np.loadtxt(fname, delimiter=',', skiprows=1)

		# withdraw concentrations (ppb in gas, 
		# # molecules/cm3 in particle and wall)
		fname = str(outd_dir + 
		'/nudged_concentrations_all_components_all_times_gas_particle_wall')
		y = np.loadtxt(fname, delimiter=',', skiprows=1)

		# convert gas-phase concentrations from ppb to molecules/cm3
		y[:, 0:nc] = y[:, 0:nc]*(np.array((Cfactor)).reshape(-1, 1))

		# get number of particle size bins, remembering to subtract
		# the gas phase
		nsb = int(x.shape[1])


		# in case you want to see SOPM mass fraction outdoors
		# prepare array for holding indices of SOPM components
		#SOPMi = np.zeros((1, nc0))

		# get the indices of secondary organic components
		#for compi in range(nc0):
			#if rel_SMILES[compi].count('C')+rel_SMILES[compi].count('c')>0:
				#if 'pri_org' not in comp_names[compi]:
					#SOPMi[0, compi] = 1

		# convert to boolean array and tile over times
		#SOPMi = np.squeeze(SOPMi == 1)

		# prepare to hold mole fractions of outdoor components
		seedx_out = np.zeros((len(t_array), nc0*nsb))

		# get particle-phase mole fractions of outdoor components 
		# at all 
		# times per size bin and ensure they sum to 1
		for i in range(0, nsb):
			seedx_out[:, nc0*i:nc0*(i+1)] = y[:, nc0*(i+1):nc0*(i+2)]/(
				np.sum(y[:, nc0*(i+1):nc0*(i+2)], axis=1).reshape(-1, 1))

			# in case you want to see SOPM mass fraction outdoors
			# estimate and print out the mass fraction of outdoor PM2.5
			# that is SOPM
			#if (i == 0): # first size bin range is PM2.5
				#mass_frac_all_comp = seedx_out[
				#	:, nc0*i:nc0*(i+1)]*(y_MM.reshape(1, -1))

				# mass fraction of just SOPM components
				#mf_SOPM = np.zeros((y.shape[0], nc0))

				# zero non-SOPM components
				#for it in range(y.shape[0]):
				#	mf_SOPM[it, SOPMi] = mass_frac_all_comp[it, SOPMi]
				
				# ensure mass fractions sum to 1
				#mass_frac_SOPM = (np.sum(mf_SOPM, axis=1)/np.sum(
				#	mass_frac_all_comp, axis=1))
				# average mass fraction of SOPM over time 
				#mass_frac_SOPM = np.sum(mass_frac_SOPM)/len(mass_frac_SOPM)
				#print(str('outdoor mass fraction of PM2.5 that ' + 
				#'is SOPM averaged over time: ' + str(mass_frac_SOPM)))

		# isolate just gas-phase concentrations (molecules/cm3)
		y_g = y[:, 0:nc]

		# zero the concentration of H2O, as this set inside PyCHAM
		# based on the RH model variables
		y_g[:, comp_names.index('H2O')] = 0.

		# withdraw number-size distributions (# particles/cm3 (air))
		fname = str(outd_dir + '/particle_number_concentration_wet')
		Nwet = np.loadtxt(fname, delimiter=',', skiprows=1)
		if (nsb == 1): # if just one size bin, ensure two dimensions
			Nwet = Nwet.reshape(-1, 1)


		# generate an array with air change rates aligned with
		# outdoor concentration times (/s)
		acr_t_array = np.zeros((len(t_array)))
		for ti in range(len(t_array)):
			acr_t_array[ti] = acr[np.sum(acr_times>=t_array[ti])-1]


		# prepare to hold names of particle-phase components
		# emitted indoors
		ind_emi_PM_comp_names = []

		# remember indices of gas-phase components emitted 
		# indoors relative to all outdoor and indoor components
		ind_comp_indx = []

		# remember indices of gas-phase components emitted 
		# indoors relative to just indoor components
		ind_emi_val_g_indx = []

		# remember original suffix for any particulate matter
		# size bins
		pm_suffix = []

		# hold indices of PM components in in_emi_comp_names
		# split by size bin
		pm_indx = np.zeros((nsb, 0))

		# count on indoor components
		ind_cnt = -1

		# count on size bins
		sbi = 0

		# prepare to hold densities and molar masses of
		# particle-phase components emitted indoors
		y_dens_ind = []
		y_MM_ind = []
		# track indoor emitted components who's 
		# properties have already been found
		prop_list = []

		# ensure original indoor emission
		# component names stays the same
		ind_emi_comp_names = np.zeros((
			len(ind_emi_comp_names0))).astype('str')
		ind_emi_comp_names[:] = ind_emi_comp_names0[:]

		# check for any components present in indoor
		# emissions that are not in outdoor concentrations 
		# and append to comp_names
		for ind_compi in ind_emi_comp_names:

			# count on indoor components
			ind_cnt += 1
		
			if 'PM' in ind_compi or 'pm' in ind_compi:
			
				# get index of final _
				us_indx = -ind_compi[-1::-1].index('_')-1

				# remember particle size bin suffix
				pm_suffix.append(ind_compi[us_indx::])

				# remove particle size bin from string
				ind_compi = ind_compi[0:us_indx]
			
				# if this particle-phase component not yet
				# contained in list of indoor particle-
				# phase components
				if ind_compi not in ind_emi_PM_comp_names:
					ind_emi_PM_comp_names.append(ind_compi)
					pm_comp_indx = -1
				else:
					pm_comp_indx = ind_emi_PM_comp_names.index(ind_compi)
			
				# check whether component is in outdoor components
				if ind_compi in comp_names and ind_compi not in prop_list:
					y_dens_ind.append(dens[comp_names.index(ind_compi)])
					y_MM_ind.append(y_MM[comp_names.index(ind_compi)])
					prop_list.append(ind_compi)

				# in case indoor component name has 'in' appended to end
				if (ind_compi[-2::] == 'in' and 
					ind_compi[0:-2] in comp_names and 
					ind_compi not in prop_list):
						y_dens_ind.append(
						dens[comp_names.index(ind_compi[0:-2])])
						y_MM_ind.append(y_MM[
						comp_names.index(ind_compi[0:-2])])
						prop_list.append(ind_compi)

				# first addition of component to pm_indx (further
				# additions possible below)
				if (len(pm_suffix) == 1):
					pm_indx = np.concatenate((pm_indx, 
					np.zeros((nsb, 1))), axis=1)
					pm_indx[sbi, pm_comp_indx] = int(ind_cnt)
				else:
					# if moving up a size bin
					if pm_suffix[-2] != pm_suffix[-1]:
						sbi += 1
						pm_indx[sbi, pm_comp_indx] = int(ind_cnt)
					# addition of more components to pm_indx
					# if on same size bin
					else:
						if (sbi == 0): # if still on first size bin
							pm_indx = np.concatenate((pm_indx, 
							np.zeros((nsb, 1))), axis=1)
					
						pm_indx[sbi, pm_comp_indx] = int(ind_cnt)

			
				continue	

			# if a gas-phase component and it is new to 
			# comp_names (which is originally based on outdoor components)
			if ind_compi not in comp_names:
				comp_names.append(ind_compi)
			
			# if a gas-phase component, then remember
			# index relative to all outdoor and indoor components
			ind_comp_indx.append(comp_names.index(ind_compi))
			# also remember index relative to just indoor components
			ind_emi_val_g_indx.append(ind_cnt)

		# ensure integer type
		pm_indx = pm_indx.astype('int')
	
		# keep just the unique PM suffixes in the order 
		# they were read in
		ic = 1 # count on unique components
	
		for i in range(1, len(pm_suffix)):
			if pm_suffix[ic] == pm_suffix[ic-1]:
				pm_suffix = pm_suffix[0:ic] + pm_suffix[ic+1::]
			else:
				ic += 1 # count on unique components

		# for the outdoor particle-phase, identify components present 
		# across any particle size bins
		# sum components over all times
		y_sum = (np.sum(y[:, nc0:nc0*(nsb+1)], axis=0)).reshape(-1, 1)
		# rearrange so components in columns and size bins in rows
		y_sum = y_sum.reshape(nsb, nc0)
		# sum over size bins
		y_sum = (np.sum(y_sum, axis = 0)).reshape(-1, 1)
		# get index of components present in any size bin
		pres_indx = (y_sum[:, 0] > 0.)
		# get names of these components, note that to this list
		# unique particle-phase components from indoors will be
		# added below
		part_comps = np.array((comp_names[0:nc0]))[pres_indx]

		# prepare to hold indices of indoor particulate
		# matter components relative to cont_infl
		cont_infl_index_ind_pp = []

		# initial flag for indoor emission of
		# particle number concentration
		ind_pconc_flag = 0
	
		# remember number of components contained in outdoor
		# particle-phase components, this used below to create
		# the cont_infl index for outdoor particle phase
		# components
		len_part_comps_out = len(part_comps)

		# check for particle-phase components present in
		# indoor emissions that are absent in outdoor 
		# (either absent in both outdoor particle-phase
		# and gas-phase in the case of appending to
		# comp_names, or absent in just outdoor 
		# particle-phase in the case of appending to part_comps)
		# component names, note this has to happen after 
		# the loop above that appends gas-phase components
		# from the indoors that are not in the outdoor
		# simulation to comp_names, so that names are in
		# order
		for ie_PM_name in ind_emi_PM_comp_names:
		
			# we are looking for actual components, not
			# particle descriptors like pconc or mean_rad 
			# (these dealt with in next conditions)
			if ('pconc' not in ie_PM_name and 
				'mean_rad' not in ie_PM_name):
				if (ie_PM_name not in comp_names):
					# append to list of all component names
					comp_names.append(ie_PM_name)
					# budge indoor particle-phase component
					# indices up by one, because these should be
					# after comp_names
					cont_infl_index_ind_pp = [x+1 for x 
					in cont_infl_index_ind_pp]
				if (ie_PM_name not in part_comps):
					part_comps = np.concatenate((
					part_comps, np.array((ie_PM_name)).reshape(1)), axis=0)
		
				# index of this component in the particle phase
				# relative to cont_infl
				cont_infl_index_ind_pp.append(
				len(comp_names)+np.where(part_comps==ie_PM_name)[0])
		
		# remember the index of cont_infl (values for cont_infl
		# set below), for non-zero outdoor particle-phase 
		# components, size bins are in columns and are filled
		# in below
		cont_infl_index_out_pp = (np.arange(
			len(comp_names), (len(comp_names)+len_part_comps_out), 
			1).astype(int)).reshape(-1, 1)

		# convert to array, with size bins in columns,
		# with size bins filled in below
		cont_infl_index_ind_pp = np.array((
		cont_infl_index_ind_pp)).reshape(-1, 1)
	
		# index of pconc in cont_infl, with further size bins
		# dealt with below
		cont_infl_index_pconc = np.array((
			len(comp_names)+len(part_comps)+1)).reshape(1)
	
		# index of pconc and mean_rad in first particle size bin, 
		# with further size bins dealt with below
		cont_infl_index_mean_rad = np.array((
			len(comp_names)+len(part_comps)+2)).reshape(1)

		# update the number of components
		nc = len(comp_names)

		# prepare to hold particle components names per size bin	
		part_comps_full_str = []

		# index array to point to indices of 
		# particle-phase components
		seedx_indx = np.zeros((len(part_comps), nsb))
	
		# prepare to hold total particle-phase 
		# molecular concentration per size bin
		# from outdoors influxing indoors, times in rows, size bins
		# in columns
		tot_pp_out = np.zeros((len(t_array), nsb))

		# prepare to hold total (summed across components) 
		# particle-phase molecular concentration 
		# influxed from 
		# indoor activities (molecules/cm3/s), size bins in rows,
		# activities in columns
		tot_pp_in = np.zeros((nsb, ind_emi_val.shape[1]))

		# prepare to hold volumes of particles from outdoors
		# per size bin (um3)
		tot_v_out_sbi = np.zeros((len(t_array), nsb))

		# prepare to hold volumes of particles from indoors
		# per activity and per size bin (um3)
		tot_v_in_sbi = np.zeros((nsb, ind_emi_val.shape[1]))

		# prepare to store particle number concentration
		# influx from outdoors, with times in rows and
		# size bins in columns
		Nwet_out_store = np.zeros((len(t_array), nsb))

		# prepare to store particle number concentration influxed from
		# indoor activities in this size bin (particles/cm3/s)
		# with size bins in rows and activities in columns
		ind_N_sbi_store = np.zeros((nsb, ind_emi_val.shape[1]))

		# loop through size bins
		for sbi in range(nsb):

			# format for particle-phase components in a given
			# size bin for the cont_infl file
			sb_str = str('_pm' + str(sbi) + '_seedx')

			# prepare to hold abundances of particle-phase 
			# indoor-emitted components in just this size bin
			ind_emi_sbi = np.zeros((0, ind_emi_val.shape[1]))

			# change PM suffix to be consistent with PM suffix from
			# outdoor simulation
			# loop through particle-phase components
			# emitted indoors
			for ind_pmi in range(len(ind_emi_comp_names[pm_indx[sbi, :]])):

				# get name of component now
				ind_comp_now = ind_emi_comp_names[pm_indx[sbi, ind_pmi]]

				# name of indoor emitted component 
				if ('pconc' not in ind_comp_now and 
					'mean_rad' not in ind_comp_now):
					ind_emi_comp_names[
					pm_indx[sbi, ind_pmi]] = ind_comp_now.replace(
					pm_suffix[sbi], sb_str)

					# keep the mole fraction of this indoor emitted 
					# particle-phase component
					# in this size bin for all activities
					ind_emi_sbi = np.concatenate((
					ind_emi_sbi, 
					ind_emi_val[pm_indx[sbi, ind_pmi]].reshape(
					1, -1)), axis= 0)
					
				# the indoor emitted particle number concentration rate 
				# in this
				# size bin (particles/cm3/s) per activity, 
				# note that the particle
				# number concentration is divided by the total volume of the
				# envelope (cm3) (e.g. household) to convert from particles/s to
				# particles/cm3/s, i.e. accounting for dilution throughout
				# the envelope 
				if 'pconc' in ind_comp_now:
					ind_N_sbi = np.array((
					ind_emi_val[
					ind_emi_comp_names.tolist().index(ind_comp_now), :
					])).reshape(1, -1)/env_vol
					

				# note the indoor emitted particle radius in this size bin (um)
				# per activity
				if 'mean_rad' in ind_comp_now:
					ind_x_sbi = np.array((
					ind_emi_val[
					ind_emi_comp_names.tolist().index(
					ind_comp_now), :])).reshape(1, -1)

			# fill in later size bins of continuous influx array 
			# index for outdoor and indoor particle-phase 
			# components, pconc and mean_rad, add +2 to allow 
			# for pconc and mean_rad,
			# per size bin
			if (sbi > 0):
				cont_infl_index_out_pp = np.concatenate((
				cont_infl_index_out_pp, (cont_infl_index_out_pp[:, -1]+
				len(part_comps)+2).reshape(-1, 1)), axis = 1)
				cont_infl_index_ind_pp = np.concatenate((
				cont_infl_index_ind_pp, (cont_infl_index_ind_pp[:, -1]+
				len(part_comps)+2).reshape(-1, 1)), axis = 1)
				cont_infl_index_pconc = np.concatenate((
				cont_infl_index_pconc, (cont_infl_index_pconc[-1]+
				len(part_comps)+2).reshape(1)), axis=0)
				cont_infl_index_mean_rad = np.concatenate((
				cont_infl_index_mean_rad, (cont_infl_index_mean_rad[-1]+
				len(part_comps)+2).reshape(1)), axis=0)

			# ensure mole fractions in particle-phase of 
			# indoor-emitted components sum to 1 per activity
			# sum of component abundances per activity
			ind_emi_sbi_sum = (np.sum(ind_emi_sbi, axis=0)).reshape(1, -1)
			# indices of where abundances are greater than 0
			pos_indx = ind_emi_sbi_sum[0, :]>0.

			# sum mole fractions to 1 per activity
			ind_emi_sbi[:, pos_indx] = ind_emi_sbi[
			:, pos_indx]/ind_emi_sbi_sum[0, pos_indx]

			# record mole fractions of components (rows) per activity 
			# (columns) per size bins (columns)
			if (sbi == 0): # set up matrix
				ind_emi_all_sb = np.zeros((ind_emi_sbi.shape[0], 
					ind_emi_sbi.shape[1], nsb))

			# record mole fractions for this size bin
			ind_emi_all_sb[:, :, sbi] = ind_emi_sbi
			
			# label particle components with particle size bin
			# number
			sb_str = [s + sb_str for s in part_comps]

			# append these components onto component names
			part_comps_full_str += sb_str

			# append particle concentration onto component
			# names (particles/cm3/s)
			part_comps_full_str += [str('pconc_pm' + str(sbi))]

			# append particle mean radius onto component 
			# names (um)
			part_comps_full_str += [str('mean_rad_pm' + str(sbi))]

			# hold the outdoor concentrations of all 
			# components in this size bin, including 
			# filler values for those components emitted 
			# indoors, times in rows
			y_out_sbi = np.zeros((len(t_array), nc))

			# outdoor concentrations of outdoor components in 
			# this size bin (molecules/cm3)
			y_out_sbi[:, 0:nc0] = y[:, nc0*(sbi+1):nc0*(sbi+2)]

			# prepare for particle-phase component concentration
			# calculation --------------------------------------

			# spread outdoor particle radius in this size bin
			# across activities
			x_out = np.tile(x[:, sbi].reshape(-1, 1), (1, ind_x_sbi.shape[1]))
			# indoor particle radius shape
			ind_x_sbi = ind_x_sbi.reshape(1, -1)

			# convert outdoor particle number concentration
			# to influx rate of outdoor particles 
			# (particles/cm3/s)
			Nwet_sbi = Nwet[:, sbi]*acr_t_array
			# store outdoor particle number concentration
			# influx rate, with times in rows and size 
			# bins in columns
			Nwet_out_store[:, sbi] = Nwet_sbi

			# spread outdoor particle number concentration
			# in this size bin across activities
			Nwet_out = np.tile(Nwet_sbi.reshape(-1, 1), (1, ind_x_sbi.shape[1]))

			# indoor particle number concentration shape 
			# (activities in columns)
			ind_N_sbi = ind_N_sbi.reshape(1, -1)

			# store particle number concentration influxed from
			# indoor activities in this size bin (particles/cm3/s)
			# with size bins in rows and activities in columns
			ind_N_sbi_store[sbi, :] = ind_N_sbi

			# total outdoor volume of influxed particles in this size bin
			# um3/s, times in rows
			tot_v_out_sbi[:, sbi] = ((4./3.)*np.pi*(x_out**3.*Nwet_out))[:, 0]
			# total indoor volume of influxed particles in this size bin
			# um3/s, activites in columns
			tot_v_in_sbi[sbi, :] = (4./3.)*np.pi*(ind_x_sbi**3.*ind_N_sbi)

			# outdoor seed component mole fractions 
			# in this size bin at all times
			seedx_out_sbi = seedx_out[:, nc0*sbi:nc0*(sbi+1)]
		
			# arithmetic mean density of outdoor particles (g/cm3)
			out_dens = np.sum(seedx_out_sbi[:, 0:nc0]*dens.reshape(1, -1), axis=1)
			# arithmetic mean molar mass of outdoor particles (g/mol)
			out_MM = np.sum(seedx_out_sbi[:, 0:nc0]*y_MM.reshape(1, -1), axis=1)
			# arithmetic mean molar volume of outdoor particles (um3/molecules)
			# note 1e12 converts cm3 to um3 and /Avogadro's constant converts
			# mol to molecules
			out_MV = (out_dens/out_MM)*1e12/si.N_A

			# total (summed across components) particle-phase 
			# molecular concentration influxing indoors
			# from outdoors (molecules/cm3/s)
			# in this size bin (columns) at all times (rows)
			tot_pp_out[:, sbi] = np.squeeze(
				tot_v_out_sbi[:, sbi].reshape(-1, 1)/out_MV.reshape(-1, 1))

			# indoor seed component mole fractions 
			# in this size bin, components in rows
			# activities in columns
			seedx_in_sbi = ind_emi_sbi
			# arithmetic mean density and molar mass 
			# of indoor particles (g/cm3) for each activity, 
			# note components are in rows and activities are in
			# columns
			in_dens = np.sum(seedx_in_sbi*np.array((
			y_dens_ind)).reshape(-1, 1), axis=0)
			# arithmetic mean molar mass of indoor particles (g/mol)
			in_MM = np.sum(seedx_in_sbi*
				np.array((y_MM_ind)).reshape(-1, 1), axis=0)
			# arithmetic mean molar volume of indoor particles (um3/molecule)
			# note 1e12 converts cm3 to um3 (components in rows, activities
			# in columns) and /Avogadro's constant converts
			# mol to molecules
			# activities with non-zero particle-phase component abundance
			pos_indx = np.sum(seedx_in_sbi, axis=0)>0.
			in_MV = np.zeros((in_dens.shape[0]))
			# mean molar volume for each activity (um3/molecule)
			in_MV[pos_indx] = (in_dens[pos_indx]/in_MM[pos_indx])*1e12/si.N_A

			# total (summed across components) particle-phase molecular 
			# concentration influx from each
			# indoor activities (molecules/cm3/s), size bins in rows,
			# activities in columns
			tot_pp_in[sbi, pos_indx] = (tot_v_in_sbi[sbi, pos_indx]/
				in_MV[pos_indx].reshape(1, -1))

	

		# ensure integer
		seedx_indx = seedx_indx.astype('int')

		# append particle-phase component names onto all component
		# names
		comp_names += part_comps_full_str

		# withdraw indices of peroxy and alkoxy radicals
		# path
		load_path = str(outd_dir + '/organic_peroxy_radical_index.npy')
		RO2i= (np.load(load_path, allow_pickle=True)).tolist()

		# path
		load_path = str(outd_dir + '/organic_alkoxy_radical_index.npy')
		ROi = (np.load(load_path, allow_pickle=True)).tolist()

		# set penetration factors for components entering
		# from outdoors in the gas phase
		pfs_g = np.ones((nc0))
		# set penetration factor to zero for components too short-lived to
		# cross building envelope:
		# OH, HO2, NO3, RO2, RO
		pfs_g[comp_names.index('OH')] = 0.
		pfs_g[comp_names.index('HO2')] = 0.
		pfs_g[comp_names.index('NO3')] = 0.
		pfs_g[RO2i] = 0.
		pfs_g[ROi] = 0.

		# prepare array to hold continuous influxes (molecules/cm3/s),
		# note that times are in columns and provide input 
		# throughout the day. Components are in rows, and
		# will include all components (present outdoor and indoor)
		# for the gas-phase, whilst for the particle-phase, only
		# those with non-zero values from outdoors and all
		# components mentioned indoors are included per size bin, in addition
		# to the pconc and mean_rad per size bin (as contained in 
		# comp_names). Note the final +1 in row length to allow a row for
		# times through simulation
		cont_infl = np.zeros((len(comp_names)+1,  int(24.*3600./300.)))

		# apply time through simulation (s) to continuous influx matrix
		cont_infl[0, :] = np.arange(0, 8.64e4, 300.)[:]

		# apply the air change rate to continuous influx array (/s)
		# keep count on air change rate times
		acrti = 0
		# keep count on outdoor concentration times
		octi = 0
		# keep count on activity times
		actti = 0

		# loop through times
		for ti in range(cont_infl.shape[1]):

			# if not already at final activity time
			if (actti+1 < len(act_matrix[0, :])):
				if (cont_infl[0, ti] == act_matrix[0, :][actti+1]):
					actti += 1
		
			# if not already at final air change rate
			if (acrti+1 < len(acr_times)):
				if (cont_infl[0, ti] == acr_times[acrti+1]):
					acrti += 1

			# if not already at final time for 
			# outdoor concentrations
			if (octi+1 < len(t_array)):
				if (cont_infl[0, ti] == t_array[octi+1]):
					octi += 1
	
			# rate of gas-phase component influx (molecules/cm3/s), 
			# accounting for outdoor concentration, air change rate and 
			# penetration factor. Further below is particle-phase 
			# influx from outside and influxes from indoor activities and
			# the total number concentration and mean radius of particles
			# per size bin
			cont_infl[1:(nc0+1), ti] = (y_g[octi, :]*acr[acrti]*pfs_g)

			# get the activity types now
			act_types_now = act_matrix_act_order[act_matrix[1::, actti] == 1]

			# prepare to hold indices of activities happening now, in 
			# preparation for particle-phase part
			all_ind_act_indx = []

			# loop through activities happening now to get continuous
			# influx of gases from indoor activities (molecules/cm3/s)
			for acti in act_types_now:

				# indices of activities happening now
				act_indx = comp_inf_col_head.index(acti)

				# store indicies of activities happening now
				all_ind_act_indx.append(act_indx)
	
				# indoor influx rate of gas-phase components (molecules/cm3/s)
				# from indoor activities (summing within this loop across 
				# activities), note that the divide by the volume of the 
				# envelope being emitted into (cm3) and therefore diluted
				# by, converts units from molecules/s to molecules/cm3/s
				cont_infl[1::, ti][ind_comp_indx] += ind_emi_val[
					ind_emi_val_g_indx, act_indx]/env_vol

			# particle-phase calculation for cont_infl starts ------------

			# loop through size bins
			for sbi in range(nsb):

				# hold total (summed across components) molecular 
				# influx rate of particle-phase components (molecules/cm3/s)
				# in this size bin at this time,
				# starting with outdoor influx rate 
				tot_mol_inf_now = tot_pp_out[octi, sbi].reshape(1, -1)

				# prepare to hold total (summed across activities) particle 
				# number concentration
				# from indoor activities in this size bin now (particles/cm3/s)
				ind_N_sbi_now = 0

				# loop through activities happening now
				for acti in act_types_now:
			
					# hold total (summed across activities) particle 
					# number concentration
					# from indoor activities in this size bin 
					# now (particles/cm3/s)
					ind_N_sbi_now += ind_N_sbi_store[
					sbi, comp_inf_col_head.index(acti)]

				
					# total particle-phase molecular concentration 
					# influx from this 
					# indoor activity in this size bin (molecules/cm3/s)
					# appended to influxes from outdoors and other
					# indoor sources
					tot_mol_inf_now = np.concatenate((tot_mol_inf_now, 
					tot_pp_in[sbi, comp_inf_col_head.index(acti)].reshape(
					1, 1)), axis=0)

				# sum (summed across outdoor influx and 
				# indoor activities) of all particle-phase 
				# molecular concentration 
				# influxes now (molecules/cm3/s)
				sum_tot_mol_inf_now = np.sum(tot_mol_inf_now, axis=0)

				# mole fractions of particle-phase component from outdoors
				mf_pp_out = tot_pp_out[octi, sbi]/sum_tot_mol_inf_now

				# mole fractions of particle-phase components from individual
				# indoor sources (sources in columns)
				mf_pp_in = (tot_pp_in[sbi, all_ind_act_indx]/
					sum_tot_mol_inf_now).reshape(1, -1)

				# outdoor individual seed component mole fractions 
				# in this size bin at this time
				seedx_out_sbi = (seedx_out[octi, nc0*sbi:nc0*(sbi+1)].reshape(
					1, -1)*mf_pp_out)
			
				# individual seed component mole fractions
				# in this size bin at this time for 
				# individual indoor sources, note that mf_pp_in should
				# have sources in columns
				seedx_ind_sbi = ind_emi_all_sb[
				:, all_ind_act_indx, sbi]*mf_pp_in
				
				# mean density of influxed particle-phase 
				# components from outside (g/cm3)
				dens_out = (np.sum(seedx_out_sbi*dens.reshape(1, -1), axis=1))
				
				# mean molar mass of influxed particle-phase 
				# components from inside activity (g/mol), 
				# note sum over indoor activities followed by sum 
				# over components
				dens_ind = (np.sum((np.sum(
				seedx_ind_sbi, axis = 1))*np.array((y_dens_ind)), axis = 0))
			
				# mean molar mass of influxed particle-phase 
				# components from outside (g/mol)
				MM_out = (np.sum(seedx_out_sbi*y_MM.reshape(1, -1), axis=1))
				# mean molar mass of influxed particle-phase 
				# components from inside activity (g/mol), 
				# note sum over indoor activities followed by sum 
				# over components
				MM_ind = (np.sum(np.sum(
				seedx_ind_sbi, axis = 1)*np.array((y_MM_ind)), axis=0))


				# arithmetic mean density (g/cm3) over all sources
				dens_all = dens_out+dens_ind
				# arithmetic mean molar mass (g/mol) over all sources
				MM_all = MM_out+MM_ind
			
				# arithmetic mean molar volume (cm3/mol) for influxes
				# from all sources in this size bin at this time
				MV_all = MM_all/dens_all
			
				# convert molar volume from cm3/mol to um3/molecule
				MV_all = MV_all*1.e12/si.N_A

				# seed mole fraction 
				# influx from outdoors at this time, for this 
				# size bin, note the 1+ in
				# the row index to allow for times
				cont_infl[1+cont_infl_index_out_pp[:, sbi], 
				ti] += seedx_out_sbi[0, pres_indx]			

				# molecular concentration influx from indoors sources
				# (molecules/cm3/s)
				# zero volume of any indoor sources that
				# are not occurring at this time
				tot_v_ind_sbi_now = np.zeros((tot_v_in_sbi.shape[1]))
				tot_v_ind_sbi_now[all_ind_act_indx] = tot_v_in_sbi[
				sbi, all_ind_act_indx]
				
				# particle-phase seed mole fraction 
				# indoors at this time in this size bin, note that
				# mole fractions per indoor
				# source in this size bin is summed over indoor 
				# sources, note that 1+ allows for times in first
				# row of cont_infl
				cont_infl[1+cont_infl_index_ind_pp[:, sbi], ti] += (np.sum(
					seedx_ind_sbi, axis=1))

				# sum of influxing particle number concentration from
				# outdoors and indoors (particles/cm3/s) in this size bin
				# at this time
				pconc_sum = Nwet_out_store[octi, sbi]+ind_N_sbi_now

				# store particle number concentrations
				cont_infl[cont_infl_index_pconc[sbi], ti] = pconc_sum
			
				# total volume of all particles from all sources in this
				# size bin now, note the sum over indoor activities
				tot_v_all_sbi_now = tot_v_out_sbi[
					octi, sbi]+sum(tot_v_ind_sbi_now)

				# radius of this single particle assuming spherical shape (um)
				rad = ((tot_v_all_sbi_now/pconc_sum)*(3./4.)/np.pi)**(1./3.)
				# record
				cont_infl[cont_infl_index_mean_rad[sbi], ti] = rad

		# write the continuous influx file (guide: 
		# https://www.geeksforgeeks.org/python-writing-excel-file-
		# using-openpyxl-module/)
	
		# open new blank workbook
		wb = openpyxl.Workbook()

		# get active sheet
		sheet = wb.active 

		# set sheet title
		sheet.title = "cont_infl"

		# loop through rows and columns to write out data
		# note that ir starts at 1 as this how row
		# indexing works in openpyxl, and 2 is added to
		# length of comp_names which comprises 1 due to
		# the openpyxl indexing and 1 due to the need
		# for a header in the first row
		for ir in range(1, len(comp_names)+2):
			for ic in range(1, cont_infl.shape[1]+2):
				ele = sheet.cell(row = ir, column = ic)
				if (ic == 1): # first column 
					if (ir == 1): # first row of first column
						ele.value = "molec/cm3/s"
					else: # subsequent rows, first column
						ele.value = comp_names[ir-2]
				else: # later columns
					ele.value = cont_infl[ir-1, ic-2] 
				

		# save workbook
		wb.save(cont_infl_path) 

		# end of continuous influx section -----
		
	return()

# call function
mod_var_setup(path_there, chem_sch_name, xml_name, res_path, base_path)